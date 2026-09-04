"""
apps/audit/views.py
Journal d'audit + statistiques de complétude des dossiers
"""

from datetime import date
from django.db.models import Count, Q
from django.http import HttpResponse
from rest_framework import serializers
from rest_framework.response import Response
from rest_framework.views import APIView
import openpyxl
from openpyxl.styles import Font

from accounts.permissions import IsAdmin
from audit.models import AuditLog
from audit.stats import build_stats_detail
from employees.models import Employee, EmployeeDocument


class AuditLogSerializer(serializers.ModelSerializer):
    class Meta:
        model = AuditLog
        fields = [
            'id', 'username_snapshot', 'action',
            'target_model', 'target_label',
            'ip_address', 'timestamp', 'details'
        ]


class AuditLogListView(APIView):
    """GET /api/admin/audit-logs/?user=&action=&page=

    Visibilité par rôle : un SUPERADMIN voit tout le journal, y compris
    les actions des autres ADMIN/SUPERADMIN — c'est le seul rôle avec
    cette visibilité totale, et chacune de ses consultations est elle-même
    tracée (Action.VIEW_AUDIT_LOG) pour que ce pouvoir de surveillance
    reste surveillable. Un ADMIN ordinaire voit SES PROPRES actions **et**
    celles de tous les CONSULTANT (qu'il administre et dont il doit
    pouvoir vérifier l'activité de consultation — traçabilité RGPD/Loi
    18-07) — mais jamais celles d'un autre ADMIN ou d'un SUPERADMIN. Le
    filtre `user` de la query string est appliqué uniquement DANS ce
    périmètre pour un ADMIN (jamais un blanc-seing pour sortir de
    own+CONSULTANT), pas seulement caché côté UI."""
    permission_classes = [IsAdmin]

    def get(self, request):
        qs = AuditLog.objects.select_related('user').order_by('-timestamp')

        # Filtres
        user_q = request.query_params.get('user')
        action = request.query_params.get('action')
        target = request.query_params.get('target')

        if request.user.is_superadmin:
            if user_q:
                qs = qs.filter(username_snapshot__icontains=user_q)
        else:
            qs = qs.filter(
                Q(user=request.user) | Q(user__role='CONSULTANT')
            )
            if user_q:
                qs = qs.filter(username_snapshot__icontains=user_q)

        if action:
            qs = qs.filter(action=action)
        if target:
            qs = qs.filter(target_label__icontains=target)

        # Pagination manuelle (50 par page)
        page = int(request.query_params.get('page', 1))
        size = 50
        total = qs.count()
        qs = qs[(page - 1) * size: page * size]

        if request.user.is_superadmin:
            AuditLog.log(
                request, AuditLog.Action.VIEW_AUDIT_LOG,
                details={'user': user_q, 'action': action, 'target': target, 'page': page},
            )

        return Response({
            'total': total,
            'page': page,
            'total_pages': (total // size) + 1,
            'scope': 'all' if request.user.is_superadmin else 'own_and_consultants',
            'results': AuditLogSerializer(qs, many=True).data,
        })


class AdminStatsView(APIView):
    permission_classes = [IsAdmin]

    def get(self, request):
        from employees.models import TypeDocument
        
        total_emp = Employee.objects.filter(statut='actif').count()

        # Complétude par type — dynamique depuis la BDD. Les catégories
        # (ex. "ETAT CIVIL") ne sont jamais uploadables directement — les
        # exclure évite une ligne "0/51 (0%)" trompeuse ; seuls les
        # sous-types réels (feuilles) comptent.
        completude = {}
        for t in TypeDocument.objects.filter(
            is_active=True, sous_types__isnull=True
        ).select_related('parent').order_by('ordre', 'nom'):
            nb = EmployeeDocument.objects.filter(
                type_doc=t, is_active=True
            ).values('employee').distinct().count()
            completude[t.code] = {
                'label': t.nom,
                'parent_nom': t.parent.nom if t.parent_id else None,
                'nb_employes': nb,
                'pourcentage': round(nb / total_emp * 100, 1) if total_emp else 0,
                'required': t.obligatoire,
            }

        # Dossiers complets — employés qui ont tous les types obligatoires
        types_obligatoires = TypeDocument.objects.filter(
            obligatoire=True, is_active=True, sous_types__isnull=True
        )
        emp_complets = Employee.objects.filter(statut='actif')
        for t in types_obligatoires:
            emp_complets = emp_complets.filter(
                documents__type_doc=t,
                documents__is_active=True
            )
        nb_complets = emp_complets.distinct().count()

        # Activité 7 jours
        from django.utils import timezone
        from datetime import timedelta
        since = timezone.now() - timedelta(days=7)
        activite = AuditLog.objects.filter(
            timestamp__gte=since
        ).values('action').annotate(count=Count('id')).order_by('-count')

        return Response({
            'employes_actifs': total_emp,
            'dossiers_complets': nb_complets,
            'taux_completude_global': round(nb_complets / total_emp * 100, 1) if total_emp else 0,
            'completude_par_type': completude,
            'activite_7_jours': list(activite),
            'total_documents': EmployeeDocument.objects.filter(is_active=True).count(),
        })


def _parse_date_param(request, name):
    raw = request.query_params.get(name)
    if not raw:
        return None
    try:
        return date.fromisoformat(raw)
    except ValueError:
        return 'invalid'


class StatsDetailView(APIView):
    """
    GET /api/reporting/stats-detail/?date_debut=&date_fin=
    Statistiques détaillées de la page /statistiques (toujours
    organisation-wide) + "Mon activité" du compte connecté, et
    "Activité par administrateur" pour un SUPERADMIN — voir
    docs/superpowers/specs/2026-09-02-page-statistiques-design.md et
    CLAUDE.md section "Page Statistiques".
    """
    permission_classes = [IsAdmin]

    def get(self, request):
        date_debut = _parse_date_param(request, 'date_debut')
        date_fin = _parse_date_param(request, 'date_fin')
        if date_debut == 'invalid' or date_fin == 'invalid':
            return Response({'error': 'Date invalide (format attendu YYYY-MM-DD).'}, status=400)
        data = build_stats_detail(date_debut, date_fin, requesting_user=request.user)
        return Response(data)


def _stats_sheet(wb, title, headers, rows):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    for col_idx in range(1, len(headers) + 1):
        ws.cell(row=1, column=col_idx).font = Font(bold=True)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(
            len(str(headers[col_idx - 1])) + 2, 14
        )


class StatsExportView(APIView):
    """
    GET /api/reporting/stats-export.xlsx/?date_debut=&date_fin=
    Export Excel des mêmes statistiques que StatsDetailView, un onglet par
    section.
    """
    permission_classes = [IsAdmin]

    def get(self, request):
        date_debut = _parse_date_param(request, 'date_debut')
        date_fin = _parse_date_param(request, 'date_fin')
        if date_debut == 'invalid' or date_fin == 'invalid':
            return Response({'error': 'Date invalide (format attendu YYYY-MM-DD).'}, status=400)
        data = build_stats_detail(date_debut, date_fin, requesting_user=request.user)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        _stats_sheet(wb, 'Indicateurs', ['Indicateur', 'Valeur', 'Variation (%)'], [
            [label, data['indicateurs'][key]['valeur'], data['indicateurs'][key]['variation_pct']]
            for key, label in [
                ('recrutements', 'Recrutements'), ('archivages', 'Archivages'),
                ('dossiers_completes', 'Dossiers complétés'),
            ]
        ])
        _stats_sheet(wb, 'Répartition organisation', ['Direction', 'Département', 'Effectif'], [
            [r['nom'], '', r['count']] for r in data['repartition_direction']
        ] + [
            [r['direction_nom'], r['nom'], r['count']] for r in data['repartition_departement']
        ])
        _stats_sheet(wb, 'Répartition profils', ['Axe', 'Valeur', 'Effectif'], [
            ['Catégorie', r['nom'], r['count']] for r in data['repartition_categorie']
        ] + [
            ['Type de contrat', r['nom'], r['count']] for r in data['repartition_type_contrat']
        ] + [
            ['Fonction', r['nom'], r['count']] for r in data['repartition_fonction']
        ])
        _stats_sheet(wb, 'Évolution', ['Mois', 'Recrutements', 'Archivages'], [
            [e['mois'], e['recrutements'], e['archivages']] for e in data['evolution_mensuelle']
        ])
        _stats_sheet(wb, 'Démographie', ['Tranche âge', 'Effectif', 'Tranche ancienneté', 'Effectif'], [
            [
                data['pyramide_age'][i]['tranche'] if i < len(data['pyramide_age']) else '',
                data['pyramide_age'][i]['count'] if i < len(data['pyramide_age']) else '',
                data['pyramide_anciennete'][i]['tranche'] if i < len(data['pyramide_anciennete']) else '',
                data['pyramide_anciennete'][i]['count'] if i < len(data['pyramide_anciennete']) else '',
            ]
            for i in range(max(len(data['pyramide_age']), len(data['pyramide_anciennete'])))
        ])
        _stats_sheet(wb, 'Échéances contrats', ['N° Contrat', 'Employé', 'Date fin', 'Jours restants'], [
            [c['numero_contrat'], c['employee_nom'], c['date_fin'], c['jours_restants']]
            for c in data['contrats_echeance']
        ])
        _stats_sheet(wb, 'Complétude', ['Direction', 'Département', 'Total', 'Complets', 'Taux (%)'], [
            [r['nom'], '', r['total'], r['complets'], r['taux']] for r in data['completude_par_direction']
        ] + [
            [r['direction_nom'], r['nom'], r['total'], r['complets'], r['taux']]
            for r in data['completude_par_departement']
        ])
        activite_headers = [
            'Administrateur', 'Employés créés', 'Employés modifiés', 'Employés archivés',
            'Documents uploadés', 'Documents supprimés', 'Documents modifiés',
        ]
        if 'activite_par_admin' in data:
            _stats_sheet(wb, 'Activité par administrateur', activite_headers, [
                [
                    a['nom_complet'], a['employes_crees'], a['employes_modifies'], a['employes_archives'],
                    a['documents_uploades'], a['documents_supprimes'], a['documents_modifies'],
                ]
                for a in data['activite_par_admin']
            ])
        else:
            a = data['mon_activite']
            _stats_sheet(wb, 'Mon activité', activite_headers, [[
                request.user.full_name, a['employes_crees'], a['employes_modifies'], a['employes_archives'],
                a['documents_uploades'], a['documents_supprimes'], a['documents_modifies'],
            ]])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"statistiques_somiz_{data['periode']['debut']}_au_{data['periode']['fin']}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)

        AuditLog.log(
            request, AuditLog.Action.EXPORT,
            details={'type': 'statistiques', 'periode': data['periode']},
        )
        return response