"""
employees/import_views.py
Import CSV des employés en masse
"""
import csv
import io
import logging
import re
import openpyxl
from django.conf import settings
from django.db import transaction
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from accounts.permissions import IsAdmin
from employees.models import (
    Employee, Direction, Pole, Departement,
    Service, Cellule, Section, Poste, TypeContrat, Categorie, Contrat,
    ChampPersonnalise, EmployeeChampValeur, Echelle, MotifArchivage,
)

logger = logging.getLogger('audit')


_EXCEL_ESCAPE_RE = re.compile(r'_x[0-9A-Fa-f]{4}_')


def _strip_excel_escapes(value):
    """Retire les échappements `_xHHHH_` qu'Excel/openpyxl laissent parfois
    tels quels dans une cellule texte contenant un caractère de contrôle
    (ex. "B+_x0001_" au lieu de "B+") — le caractère d'origine n'a aucune
    valeur pour l'utilisateur, on le supprime plutôt que de le décoder."""
    if not isinstance(value, str) or '_x' not in value:
        return value
    return _EXCEL_ESCAPE_RE.sub('', value)


def _read_rows(file):
    """Lit un fichier .csv ou .xlsx et retourne (fieldnames, liste de dict).
    Le .xlsx évite toute la classe de bugs liée au délimiteur/ré-enregistrement
    d'un CSV dans Excel (colonnes réelles, pas de texte à ré-interpréter)."""
    if file.name.lower().endswith('.xlsx'):
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return None, []
        fieldnames = [str(h).strip() if h is not None else '' for h in header_row]
        rows = []
        for raw_row in rows_iter:
            if raw_row is None or all(v is None for v in raw_row):
                continue
            row = {}
            for key, value in zip(fieldnames, raw_row):
                if not key:
                    continue
                if value is None:
                    row[key] = ''
                elif hasattr(value, 'isoformat'):  # date/datetime
                    row[key] = value.isoformat()[:10]
                else:
                    row[key] = _strip_excel_escapes(str(value))
            rows.append(row)
        return fieldnames, rows

    # CSV
    try:
        content = file.read().decode('utf-8-sig')  # utf-8-sig gère le BOM Excel
    except UnicodeDecodeError:
        file.seek(0)
        content = file.read().decode('latin-1')  # Fallback pour Excel français

    sample = content[:1024]
    delimiter = ';' if sample.count(';') > sample.count(',') else ','
    reader = csv.DictReader(io.StringIO(content), delimiter=delimiter)
    if not reader.fieldnames:
        return None, []
    rows = [
        {k: _strip_excel_escapes(v) for k, v in row.items()}
        for row in reader
    ]
    return list(reader.fieldnames), rows


def _check_csv_size(file):
    """Même limite que les uploads de documents — empêche un CSV énorme
    d'être chargé intégralement en mémoire (DoS)."""
    max_size = settings.MAX_UPLOAD_SIZE_MB * 1024 * 1024
    if file.size > max_size:
        return Response(
            {'error': f'Fichier trop volumineux. Maximum {settings.MAX_UPLOAD_SIZE_MB} Mo.'},
            status=status.HTTP_400_BAD_REQUEST
        )
    return None


class EmployeeImportView(APIView):
    """
    POST /api/employees/import/
    Importe des employés depuis un fichier CSV.
    """
    permission_classes = [IsAdmin]

    # Colonnes obligatoires
    REQUIRED_COLS = {'matricule', 'numero_contrat', 'nom', 'prenom'}

    # Colonnes optionnelles structurelles — s'y ajoutent dynamiquement les
    # codes des ChampPersonnalise actifs (voir champs_actifs dans post()) ;
    # exposées ensemble via GET /ref/champs-personnalises/ pour le frontend
    # (page Import.jsx).
    OPTIONAL_COLS = {
        'date_naissance', 'date_embauche', 'statut',
        'direction', 'departement', 'service',
        'poste', 'type_contrat', 'categorie',
    }

    def post(self, request):
        file = request.FILES.get('file')

        if not file:
            return Response(
                {'error': 'Aucun fichier fourni.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        if not (file.name.lower().endswith('.csv') or file.name.lower().endswith('.xlsx')):
            return Response(
                {'error': 'Le fichier doit être au format CSV ou Excel (.xlsx).'},
                status=status.HTTP_400_BAD_REQUEST
            )

        size_error = _check_csv_size(file)
        if size_error:
            return size_error

        try:
            fieldnames, rows = _read_rows(file)
        except Exception:
            logger.exception("Échec de la lecture du fichier d'import employés.")
            return Response(
                {'error': "Fichier illisible. Vérifiez le format (CSV ou .xlsx) et l'encodage."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Vérifier les colonnes obligatoires
        if not fieldnames:
            return Response(
                {'error': 'Fichier vide ou mal formaté.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        cols = {c.strip().lower() for c in fieldnames}
        missing = self.REQUIRED_COLS - cols
        if missing:
            return Response(
                {'error': f'Colonnes obligatoires manquantes : {", ".join(missing)}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Cache des référentiels pour éviter N requêtes BDD
        directions = {d.nom.upper(): d for d in Direction.objects.filter(is_active=True)}
        departements = {d.nom.upper(): d for d in Departement.objects.filter(is_active=True)}
        services = {s.nom.upper(): s for s in Service.objects.filter(is_active=True)}
        postes = {p.nom.upper(): p for p in Poste.objects.filter(is_active=True)}
        types_contrat = {t.nom.upper(): t for t in TypeContrat.objects.filter(is_active=True)}
        categories = {c.nom.upper(): c for c in Categorie.objects.filter(is_active=True)}

        # Champs personnalisés actifs — colonne CSV = code du champ en
        # minuscules (RIB -> 'rib'). Entièrement dynamique : un champ ajouté
        # ou supprimé dans /parametres est immédiatement pris en compte,
        # sans changement de code (voir Import.jsx qui liste ces mêmes
        # colonnes optionnelles dynamiquement).
        champs_actifs = {c.code.lower(): c for c in ChampPersonnalise.objects.filter(is_active=True, is_systeme=False)}

        # Matricules existants pour détecter les doublons
        matricules_existants = set(
            Employee.objects.values_list('matricule', flat=True)
        )

        resultats = []
        erreurs = []
        a_creer = []

        for num_ligne, row in enumerate(rows, start=2):
            # Nettoyer les clés et valeurs
            row = {k.strip().lower(): (v or '').strip() for k, v in row.items() if k}
            ligne_erreurs = []

            matricule = row.get('matricule', '').upper()
            nom = row.get('nom', '').upper()
            prenom = row.get('prenom', '').capitalize()

            # Validations
            if not matricule:
                ligne_erreurs.append("Matricule manquant")
            elif matricule in matricules_existants:
                ligne_erreurs.append(f"Matricule {matricule} déjà existant")

            if not nom:
                ligne_erreurs.append("Nom manquant")
            if not prenom:
                ligne_erreurs.append("Prénom manquant")

            # Statut
            statut = row.get('statut', 'actif').lower()
            if statut not in ['actif', 'inactif', 'archive']:
                statut = 'actif'

            # Dates
            date_naissance = row.get('date_naissance') or None
            date_embauche = row.get('date_embauche') or None

            # Référentiels — résolution par nom
            direction = directions.get(row.get('direction', '').upper())
            departement = departements.get(row.get('departement', '').upper())
            service = services.get(row.get('service', '').upper())
            # Un Service appartient toujours à un Département — si la colonne
            # "departement" du CSV est vide/absente mais qu'un Service a été
            # résolu, on aligne quand même departement/direction dessus
            # (même logique que EmployeeCreateUpdateSerializer.validate()
            # pour Cellule/Section), sinon la fiche employé perd son
            # Département alors que le Service l'affiche bien.
            if service and not departement:
                departement = service.departement
            if departement and not direction:
                direction = departement.direction
            poste = postes.get(row.get('poste', '').upper())
            type_contrat = types_contrat.get(row.get('type_contrat', '').upper())
            categorie = categories.get(row.get('categorie', '').upper())

            if ligne_erreurs:
                erreurs.append({
                    'ligne': num_ligne,
                    'matricule': matricule or '—',
                    'erreurs': ligne_erreurs,
                })
                continue

            # Ajouter au matricule cache pour éviter doublons dans le même CSV
            matricules_existants.add(matricule)

            numero_contrat = row.get('numero_contrat', '').strip()
            if not numero_contrat:
                ligne_erreurs.append("N° contrat obligatoire")
            elif not numero_contrat.isdigit():
                ligne_erreurs.append("N° contrat invalide : chiffres uniquement (ex : 024141)")

            if ligne_erreurs:
                erreurs.append({
                    'ligne': num_ligne,
                    'matricule': matricule or '—',
                    'erreurs': ligne_erreurs,
                })
                continue

            a_creer.append({
                'employee': Employee(
                    matricule=matricule,
                    nom=nom,
                    prenom=prenom,
                    date_naissance=date_naissance or None,
                    date_embauche=date_embauche or None,
                    statut=statut,
                    direction=direction,
                    departement=departement,
                    service=service,
                    poste=poste,
                    type_contrat=type_contrat,
                    categorie=categorie,
                    created_by=request.user,
                ),
                'numero_contrat': numero_contrat,
                'type_contrat_obj': type_contrat,
                'date_debut': date_embauche or None,
                'champs_valeurs': {
                    code_lower: row.get(code_lower, '').strip()
                    for code_lower in champs_actifs
                    if row.get(code_lower, '').strip()
                },
            })

            resultats.append({
                'ligne': num_ligne,
                'matricule': matricule,
                'nom': f"{prenom} {nom}",
                'statut': 'ok',
            })

        # Import en masse dans une transaction
        nb_crees = 0
        if a_creer:
            try:
                with transaction.atomic():
                    employees_objs = [item['employee'] for item in a_creer]
                    created = Employee.objects.bulk_create(employees_objs, batch_size=500)
                    nb_crees = len(created)
                    # Recharger les employés créés pour avoir les IDs (bulk_create ne retourne pas les IDs en PG < 10)
                    matricules = [e.matricule for e in created]
                    emp_map = {e.matricule: e for e in Employee.objects.filter(matricule__in=matricules)}
                    contrats_a_creer = []
                    for item in a_creer:
                        if item['numero_contrat']:
                            emp = emp_map.get(item['employee'].matricule)
                            if emp:
                                contrats_a_creer.append(Contrat(
                                    employee=emp,
                                    numero_contrat=item['numero_contrat'],
                                    type_contrat=item['type_contrat_obj'],
                                    date_debut=item['date_debut'],
                                    statut='actif',
                                    created_by=request.user,
                                ))
                    if contrats_a_creer:
                        Contrat.objects.bulk_create(contrats_a_creer, batch_size=500)

                    # Champs personnalisés (dynamiques — voir champs_actifs ci-dessus)
                    valeurs_a_creer = []
                    for item in a_creer:
                        emp = emp_map.get(item['employee'].matricule)
                        if not emp:
                            continue
                        for code_lower, valeur in item['champs_valeurs'].items():
                            champ = champs_actifs.get(code_lower)
                            if champ:
                                valeurs_a_creer.append(EmployeeChampValeur(
                                    employee=emp, champ=champ, valeur=valeur,
                                ))
                    if valeurs_a_creer:
                        EmployeeChampValeur.objects.bulk_create(valeurs_a_creer, batch_size=500)
            except Exception:
                logger.exception("Échec de l'import CSV employés.")
                return Response(
                    {'error': "Erreur lors de l'import. Vérifiez le format du fichier ou contactez un administrateur."},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )

        return Response({
            'nb_crees': nb_crees,
            'nb_erreurs': len(erreurs),
            'nb_lignes': num_ligne - 1 if 'num_ligne' in locals() else 0,
            'erreurs': erreurs[:50],  # Max 50 erreurs affichées
            'succes': resultats[:10],  # Aperçu des 10 premiers créés
        })


class EmployeeImportTemplateView(APIView):
    """
    GET /api/employees/import/template/
    Retourne un fichier .xlsx template a telecharger.

    Distribue en .xlsx (plus en .csv) : un CSV edite dans Excel puis
    re-enregistre peut perdre son delimiteur (depend des parametres
    regionaux Windows/Excel de l'admin) et retomber en une seule colonne a
    la reouverture -- un classeur Excel a des colonnes reelles et n'a pas ce
    probleme. L'import (EmployeeImportView) accepte toujours les deux
    formats, .csv et .xlsx.
    """
    permission_classes = [IsAdmin]

    def get(self, request):
        from django.http import HttpResponse

        # Colonnes des champs personnalises actifs -- dynamiques, voir
        # champs_actifs dans EmployeeImportView.post().
        champs_codes = list(
            ChampPersonnalise.objects.filter(is_active=True).values_list('code', flat=True)
        )

        headers = [
            'matricule', 'numero_contrat', 'nom', 'prenom',
            'date_naissance', 'date_embauche', 'statut',
            'direction', 'departement', 'service',
            'poste', 'type_contrat', 'categorie',
            *[c.lower() for c in champs_codes],
        ]
        exemple = [
            '024141', '024141', 'FILALI', 'Zoheir',
            '2002-03-22', '2026-01-20', 'actif',
            'Direction Generale', 'DAP', 'Service Paie',
            'Cadre', 'CDI', '',
            *(['' for _ in champs_codes]),
        ]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Employes'
        ws.append(headers)
        ws.append(exemple)
        for col_idx, header in enumerate(headers, start=1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(len(header) + 2, 14)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="template_import_employes.xlsx"'
        wb.save(response)
        return response


class ReferentielImportView(APIView):
    """
    POST /api/ref/import/{model}/
    Importe un referentiel depuis un fichier CSV ou Excel (.xlsx).
    model : directions, poles, departements, services, cellules, postes, types-contrat, categories
    """
    permission_classes = [IsAdmin]

    MODELS = {
        'directions': {
            'model': Direction,
            'required': {'nom'},
            'optional': {'code', 'description'},
        },
        'poles': {
            'model': Pole,
            'required': {'nom', 'direction'},
            'optional': {'code', 'description'},
        },
        'departements': {
            'model': Departement,
            'required': {'nom', 'direction'},
            'optional': {'code', 'description'},
        },
        'services': {
            'model': Service,
            'required': {'nom', 'departement'},
            # 'direction' optionnelle : le nom d'un Departement n'est unique
            # qu'au sein de sa Direction (unique_together), donc deux
            # Departements de Directions differentes peuvent porter le meme
            # nom. Si la colonne 'direction' est renseignee, elle sert a
            # lever cette ambiguite ; sinon le departement n'est resolu par
            # son seul nom que s'il est le seul a porter ce nom.
            'optional': {'code', 'direction', 'description'},
        },
        'cellules': {
            'model': Cellule,
            'required': {'nom'},
            # Une Cellule est rattachee a exactement une Direction OU un
            # Departement en base (Cellule.clean()), mais dans le CSV
            # 'direction' est aussi utilisee pour lever l'ambiguite du nom
            # de departement (comme pour les Services ci-dessus) : si
            # 'departement' est rempli, il devient le parent et 'direction'
            # ne sert que de desambiguisation (les deux colonnes remplies
            # simultanement ne sont donc PAS une erreur). Aucune des deux
            # colonnes n'est a elle seule obligatoire, mais l'une des deux
            # doit etre renseignee sur chaque ligne (verifie a l'execution,
            # pas dans 'required' qui ne sait exprimer que des colonnes
            # toujours obligatoires).
            'optional': {'code', 'direction', 'departement', 'description'},
        },
        'sections': {
            'model': Section,
            'required': {'nom'},
            # Memes regles que 'cellules' (voir commentaire ci-dessus) :
            # 'direction' sert a la fois de parent direct (departement vide)
            # et de desambiguisation (departement rempli).
            'optional': {'code', 'direction', 'departement', 'description'},
        },
        'postes': {
            'model': Poste,
            'required': {'nom'},
            'optional': {'code', 'description'},
        },
        'types-contrat': {
            'model': TypeContrat,
            'required': {'nom'},
            'optional': {'description'},
        },
        'categories': {
            'model': Categorie,
            'required': {'nom'},
            'optional': {'description'},
        },
        'echelles': {
            'model': Echelle,
            'required': {'nom'},
            'optional': {'description'},
        },
        'motifs-archivage': {
            'model': MotifArchivage,
            'required': {'nom'},
            'optional': {'description'},
        },
    }

    def post(self, request, model):
        if model not in self.MODELS:
            return Response(
                {'error': f'Modele inconnu : {model}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        config = self.MODELS[model]
        file = request.FILES.get('file')

        if not file:
            return Response({'error': 'Aucun fichier fourni.'}, status=400)

        if not (file.name.lower().endswith('.csv') or file.name.lower().endswith('.xlsx')):
            return Response({'error': 'Format CSV ou Excel (.xlsx) requis.'}, status=400)

        size_error = _check_csv_size(file)
        if size_error:
            return size_error

        try:
            fieldnames, rows = _read_rows(file)
        except Exception:
            logger.exception("Echec de la lecture du fichier d'import referentiel (%s).", model)
            return Response(
                {'error': "Fichier illisible. Verifiez le format (CSV ou .xlsx) et l'encodage."},
                status=400
            )

        if not fieldnames:
            return Response({'error': 'Fichier vide ou mal formate.'}, status=400)

        cols = {c.strip().lower() for c in fieldnames}
        missing = config['required'] - cols
        if missing:
            return Response(
                {'error': f'Colonnes obligatoires manquantes : {", ".join(missing)}'},
                status=400
            )

        # Cache referentiels parents
        directions_cache = {d.nom.upper(): d for d in Direction.objects.all()}
        # Departements groupes par nom (un nom n'est unique qu'au sein
        # d'une Direction, unique_together, voir models.py) : necessaire
        # pour detecter l'ambiguite quand 'direction' n'est pas precisee
        # sur une ligne de service/cellule.
        departements_par_nom = {}
        departements_par_cle = {}
        for d in Departement.objects.select_related('direction').all():
            departements_par_nom.setdefault(d.nom.upper(), []).append(d)
            departements_par_cle[(d.direction.nom.upper(), d.nom.upper())] = d

        def resoudre_departement(row, ligne_erreurs, label='Departement'):
            """Resout un Departement par 'departement' (+ 'direction'
            optionnelle pour lever l'ambiguite). Retourne None et ajoute une
            erreur a ligne_erreurs si non trouve/ambigu."""
            dept_nom = row.get('departement', '').strip().upper()
            if not dept_nom:
                return None
            dir_nom = row.get('direction', '').strip().upper()
            if dir_nom:
                dept = departements_par_cle.get((dir_nom, dept_nom))
                if not dept:
                    ligne_erreurs.append(
                        f'{label} "{row.get("departement")}" introuvable sous la direction "{row.get("direction")}"'
                    )
                return dept
            matches = departements_par_nom.get(dept_nom, [])
            if not matches:
                ligne_erreurs.append(f'{label} "{row.get("departement")}" introuvable')
                return None
            if len(matches) > 1:
                ligne_erreurs.append(
                    f'Plusieurs departements nommes "{row.get("departement")}" existent, precisez la colonne "direction"'
                )
                return None
            return matches[0]

        ModelClass = config['model']

        # Existants pour detecter les doublons : le nom est globalement
        # unique pour directions/postes/types-contrat/categories, mais
        # scope a son parent pour departements (direction+nom, unique en
        # base) et services (departement+nom, unique en base) ; pour
        # cellules (pas de contrainte unique en base), on applique la meme
        # regle scopee par coherence avec le reste du referentiel.
        if model == 'poles':
            existants = {(p.direction_id, p.nom.upper()) for p in Pole.objects.all()}
        elif model == 'departements':
            existants = {(d.direction_id, d.nom.upper()) for d in Departement.objects.all()}
        elif model == 'services':
            existants = {(s.departement_id, s.nom.upper()) for s in Service.objects.all()}
        elif model in ('cellules', 'sections'):
            existants = {
                (c.direction_id, c.departement_id, c.nom.upper())
                for c in ModelClass.objects.all()
            }
        else:
            existants = {(None, n.upper()) for n in ModelClass.objects.values_list('nom', flat=True)}

        erreurs = []
        a_creer = []
        resultats = []

        for num_ligne, row in enumerate(rows, start=2):
            row = {k.strip().lower(): (v or '').strip() for k, v in row.items() if k}
            ligne_erreurs = []

            nom = row.get('nom', '').strip()
            if not nom:
                ligne_erreurs.append("Nom manquant")

            code = row.get('code', '')
            description = row.get('description', '')

            direction = None
            departement = None

            if model in ('poles', 'departements'):
                dir_nom = row.get('direction', '').upper()
                direction = directions_cache.get(dir_nom)
                if not dir_nom:
                    ligne_erreurs.append("Direction manquante")
                elif not direction:
                    ligne_erreurs.append(f'Direction "{row.get("direction")}" introuvable')
                cle_doublon = (direction.id if direction else None, nom.upper())

            elif model == 'services':
                if not row.get('departement', '').strip():
                    ligne_erreurs.append("Departement manquant")
                departement = resoudre_departement(row, ligne_erreurs)
                cle_doublon = (departement.id if departement else None, nom.upper())

            elif model in ('cellules', 'sections'):
                # 'direction' est a la fois : (a) le parent direct de la
                # Cellule/Section quand 'departement' est vide, et (b) simplement une
                # aide de desambiguisation du departement quand 'departement'
                # est rempli (meme colonne, deux usages — voir
                # resoudre_departement) : les deux colonnes remplies en meme
                # temps n'est donc PAS une erreur, seule 'departement' vide
                # ET 'direction' vide (aucun parent du tout) en est une.
                a_departement = bool(row.get('departement', '').strip())
                a_direction = bool(row.get('direction', '').strip())
                if a_departement:
                    departement = resoudre_departement(row, ligne_erreurs, label='Departement')
                elif a_direction:
                    dir_nom = row.get('direction', '').strip().upper()
                    direction = directions_cache.get(dir_nom)
                    if not direction:
                        ligne_erreurs.append(f'Direction "{row.get("direction")}" introuvable')
                else:
                    ligne_erreurs.append('Direction ou Departement requis (exactement un des deux)')
                cle_doublon = (
                    direction.id if direction else None,
                    departement.id if departement else None,
                    nom.upper(),
                )

            else:
                cle_doublon = (None, nom.upper())

            if nom and cle_doublon in existants:
                ligne_erreurs.append(f'"{nom}" existe deja a cet emplacement')

            if ligne_erreurs:
                erreurs.append({
                    'ligne': num_ligne,
                    'nom': nom or '\u2014',
                    'erreurs': ligne_erreurs,
                })
                continue

            existants.add(cle_doublon)

            # Construire l'objet selon le modele
            kwargs = {'nom': nom, 'description': description}
            if code:
                kwargs['code'] = code
            if direction:
                kwargs['direction'] = direction
            if departement:
                kwargs['departement'] = departement

            a_creer.append(ModelClass(**kwargs))
            resultats.append({'ligne': num_ligne, 'nom': nom})

        nb_crees = 0
        if a_creer:
            try:
                with transaction.atomic():
                    ModelClass.objects.bulk_create(a_creer, batch_size=500)
                    nb_crees = len(a_creer)
            except Exception:
                logger.exception("Echec de l'import referentiel (%s).", model)
                return Response(
                    {'error': "Erreur lors de l'import. Verifiez le format du fichier ou contactez un administrateur."},
                    status=500
                )

        return Response({
            'nb_crees': nb_crees,
            'nb_erreurs': len(erreurs),
            'nb_lignes': num_ligne - 1 if 'num_ligne' in locals() else 0,
            'erreurs': erreurs[:50],
            'succes': resultats[:10],
        })


class ReferentielImportTemplateView(APIView):
    """
    GET /api/ref/import/{model}/template/
    Retourne le template .xlsx pour le modele (accepte aussi .csv en import).
    """
    permission_classes = [IsAdmin]

    TEMPLATES = {
        'directions': {
            'headers': ['nom', 'code', 'description'],
            'example': ['Direction Generale', 'DG', 'Direction principale'],
        },
        'poles': {
            'headers': ['nom', 'code', 'direction', 'description'],
            'example': ['Pole Machines Tournantes', 'PMT', 'Direction Generale', ''],
        },
        'departements': {
            'headers': ['nom', 'code', 'direction', 'description'],
            'example': ['DAP', 'DAP', 'Direction Generale', 'Departement Administration du Personnel'],
        },
        'services': {
            'headers': ['nom', 'code', 'departement', 'direction', 'description'],
            # 'direction' optionnelle — utile seulement si plusieurs
            # departements du referentiel partagent le meme nom (voir
            # ReferentielImportView.MODELS['services']).
            'example': ['Service Paie', 'SP', 'DAP', '', ''],
        },
        'cellules': {
            'headers': ['nom', 'code', 'direction', 'departement', 'description'],
            # Au moins une des deux colonnes direction/departement doit etre
            # renseignee par ligne ; si 'departement' est rempli, 'direction'
            # devient facultative et sert seulement a lever l'ambiguite si
            # plusieurs departements portent ce nom — voir
            # ReferentielImportView.MODELS['cellules'].
            'example': ['Cellule Oeuvres Sociales', 'COS', '', 'DAP', ''],
        },
        'sections': {
            'headers': ['nom', 'code', 'direction', 'departement', 'description'],
            'example': ['Section Controle Qualite', 'SCQ', '', 'DAP', ''],
        },
        'postes': {
            'headers': ['nom', 'code', 'description'],
            'example': ['Ingenieur principal', 'ING-P', ''],
        },
        'types-contrat': {
            'headers': ['nom', 'description'],
            'example': ['CDI', 'Contrat a duree indeterminee'],
        },
        'categories': {
            'headers': ['nom', 'description'],
            'example': ['Cadre', ''],
        },
        'echelles': {
            'headers': ['nom', 'description'],
            'example': ['Echelle 10', ''],
        },
        'motifs-archivage': {
            'headers': ['nom', 'description'],
            'example': ['Fin de contrat', ''],
        },
    }

    def get(self, request, model):
        from django.http import HttpResponse

        if model not in self.TEMPLATES:
            return Response({'error': f'Modele inconnu : {model}'}, status=400)

        config = self.TEMPLATES[model]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = model[:31]
        ws.append(config['headers'])
        ws.append(config['example'])
        for col_idx, header in enumerate(config['headers'], start=1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(len(header) + 2, 14)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="template_{model}.xlsx"'
        wb.save(response)
        return response