"""
apps/employees/export_views.py
Export Excel (.xlsx) des employés — ADMIN only.

Deux endpoints :
- GET /api/employees/<pk>/export/  → un employé (infos + liste des documents)
- GET /api/employees/export/       → tous les employés respectant les
  filtres de la liste /employees (mêmes query params que EmployeeListCreateView),
  une ligne par employé.

Pas de fichiers physiques inclus (voir CLAUDE.md) — uniquement les données
et, pour un employé, la liste des documents scannés (nom, type, date, taille).
"""
import openpyxl
from openpyxl.styles import Font
from django.db.models import Q, OuterRef, Subquery
from django.http import HttpResponse
from rest_framework.views import APIView

from accounts.permissions import IsAdmin
from audit.models import AuditLog
from employees.models import Employee, ChampPersonnalise, Contrat, TypeDocument
from employees.views import resolve_employee


def _champs_personnalises_actifs():
    return list(
        ChampPersonnalise.objects.filter(is_active=True, is_systeme=False).order_by('ordre', 'nom')
    )


def _types_documents_actifs():
    # Uniquement les types "feuilles" (non catégories) — voir CLAUDE.md,
    # section "Hiérarchie des types de documents" : une catégorie n'est
    # jamais elle-même uploadable, donc jamais présente/absente pour un
    # employé.
    return list(
        TypeDocument.objects.filter(is_active=True, sous_types__isnull=True).order_by('ordre', 'nom')
    )


def _employee_row_headers(champs, types_docs):
    return [
        'Matricule', 'Nom', 'Prénom', 'N° Contrat actif',
        'Date de naissance', 'Date de recrutement', 'Statut',
        'Direction', 'Département', 'Service', 'Cellule', 'Section',
        'Fonction', 'Type de contrat', 'Catégorie',
        *[c.nom for c in champs],
        *[f'Scan : {t.nom}' for t in types_docs],
    ]


def _employee_row_values(employee, champs, types_docs, numero_contrat_actif=None, valeurs_par_champ=None, types_presents=None):
    valeurs_par_champ = valeurs_par_champ or {
        v.champ_id: v.valeur for v in employee.valeurs_personnalisees.all()
    }
    if types_presents is None:
        types_presents = set(
            employee.documents.filter(is_active=True).values_list('type_doc_id', flat=True)
        )
    return [
        employee.matricule, employee.nom, employee.prenom,
        numero_contrat_actif or '',
        employee.date_naissance.isoformat() if employee.date_naissance else '',
        employee.date_embauche.isoformat() if employee.date_embauche else '',
        employee.get_statut_display() if hasattr(employee, 'get_statut_display') else employee.statut,
        employee.direction.nom if employee.direction_id else '',
        employee.departement.nom if employee.departement_id else '',
        employee.service.nom if employee.service_id else '',
        employee.cellule.nom if employee.cellule_id else '',
        employee.section.nom if employee.section_id else '',
        employee.poste.nom if employee.poste_id else '',
        employee.type_contrat.nom if employee.type_contrat_id else '',
        employee.categorie.nom if employee.categorie_id else '',
        *[valeurs_par_champ.get(c.id, '') for c in champs],
        *['Oui' if t.id in types_presents else 'Non' for t in types_docs],
    ]


def _autosize(ws, headers):
    for col_idx, header in enumerate(headers, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(len(str(header)) + 2, 14)


def _bold_header(ws, ncols):
    for col_idx in range(1, ncols + 1):
        ws.cell(row=1, column=col_idx).font = Font(bold=True)


class EmployeeExportView(APIView):
    """
    GET /api/employees/<pk>/export/
    Export Excel d'un employé : feuille "Informations" + feuille "Documents".
    """
    permission_classes = [IsAdmin]

    def get(self, request, pk):
        employee = resolve_employee(pk, Employee.objects.select_related(
            'direction', 'departement', 'service', 'cellule', 'section',
            'poste', 'type_contrat', 'categorie',
        ).prefetch_related('valeurs_personnalisees__champ'))

        champs = _champs_personnalises_actifs()
        types_docs = _types_documents_actifs()
        latest_contrat = employee.contrats.order_by('-date_debut', '-id').first()

        wb = openpyxl.Workbook()

        ws_info = wb.active
        ws_info.title = 'Informations'
        headers = _employee_row_headers(champs, types_docs)
        ws_info.append(headers)
        ws_info.append(_employee_row_values(
            employee, champs, types_docs,
            numero_contrat_actif=latest_contrat.numero_contrat if latest_contrat else None,
        ))
        _bold_header(ws_info, len(headers))
        _autosize(ws_info, headers)

        ws_docs = wb.create_sheet('Documents')
        doc_headers = ['Type de document', 'Catégorie parente', 'Nom du fichier', "Date d'upload", 'Taille (Mo)', 'Version']
        ws_docs.append(doc_headers)
        documents = employee.documents.filter(is_active=True).select_related('type_doc', 'type_doc__parent').prefetch_related('fichiers')
        for doc in documents:
            for f in doc.fichiers.filter(is_active=True).order_by('ordre'):
                ws_docs.append([
                    doc.type_doc.nom if doc.type_doc else '',
                    doc.type_doc.parent.nom if doc.type_doc and doc.type_doc.parent_id else '',
                    f.file_name,
                    f.uploaded_at.strftime('%Y-%m-%d %H:%M') if f.uploaded_at else '',
                    round((f.file_size or 0) / (1024 * 1024), 2),
                    doc.version,
                ])
        _bold_header(ws_docs, len(doc_headers))
        _autosize(ws_docs, doc_headers)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"employe_{employee.matricule}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)

        AuditLog.log(
            request, AuditLog.Action.EXPORT,
            target=employee,
            details={'type': 'employe_unique', 'matricule': employee.matricule, 'nb_documents': documents.count()}
        )
        return response


class EmployeeExportAllView(APIView):
    """
    GET /api/employees/export/
    Export Excel de tous les employés (respecte les mêmes filtres query
    params que EmployeeListCreateView — q, direction, departement, service,
    pole, cellule, section, statut), une ligne par employé.
    """
    permission_classes = [IsAdmin]

    def get(self, request):
        latest_contrat = Contrat.objects.filter(
            employee=OuterRef('pk')
        ).order_by('-date_debut', '-id')

        qs = Employee.objects.select_related(
            'direction', 'departement', 'service', 'cellule', 'section',
            'poste', 'type_contrat', 'categorie',
        ).prefetch_related('valeurs_personnalisees__champ').annotate(
            numero_contrat_actif=Subquery(latest_contrat.values('numero_contrat')[:1]),
        )

        q = request.query_params.get('q')
        direction = request.query_params.get('direction')
        dept = request.query_params.get('departement')
        service = request.query_params.get('service')
        pole = request.query_params.get('pole')
        cellule = request.query_params.get('cellule')
        section = request.query_params.get('section')
        statut = request.query_params.get('statut')
        vue = request.query_params.get('vue')

        if q:
            qs = qs.filter(
                Q(nom__icontains=q) |
                Q(prenom__icontains=q) |
                Q(matricule__icontains=q)
            )
        if direction:
            qs = qs.filter(direction=direction)
        if dept:
            qs = qs.filter(departement=dept)
        if service:
            qs = qs.filter(service=service)
        if pole:
            qs = qs.filter(departement__pole=pole)
        if cellule:
            qs = qs.filter(cellule=cellule)
        if section:
            qs = qs.filter(section=section)
        # Même comportement par défaut que EmployeeListCreateView — voir
        # CLAUDE.md section Archivage employé.
        if statut:
            qs = qs.filter(statut=statut)
        elif vue == 'archives':
            qs = qs.filter(statut__in=['inactif', 'archive', 'demobilise'])
        else:
            qs = qs.filter(statut='actif')

        qs = qs.order_by('nom', 'prenom').prefetch_related('documents')

        champs = _champs_personnalises_actifs()
        types_docs = _types_documents_actifs()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Employés'
        headers = _employee_row_headers(champs, types_docs)
        ws.append(headers)
        for employee in qs:
            valeurs = {v.champ_id: v.valeur for v in employee.valeurs_personnalisees.all()}
            types_presents = {d.type_doc_id for d in employee.documents.all() if d.is_active}
            ws.append(_employee_row_values(
                employee, champs, types_docs,
                numero_contrat_actif=employee.numero_contrat_actif,
                valeurs_par_champ=valeurs,
                types_presents=types_presents,
            ))
        _bold_header(ws, len(headers))
        _autosize(ws, headers)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="export_employes.xlsx"'
        wb.save(response)

        AuditLog.log(
            request, AuditLog.Action.EXPORT,
            details={'type': 'tous_employes', 'nb_employes': qs.count(), 'filtres': dict(request.query_params)}
        )
        return response
