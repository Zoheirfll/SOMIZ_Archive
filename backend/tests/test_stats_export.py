import pytest
import openpyxl
import io
from rest_framework.test import APIClient
from audit.models import AuditLog


def auth_client(user):
    client = APIClient()
    client.force_authenticate(user=user)
    return client


@pytest.mark.django_db
class TestStatsExportView:
    def test_admin_can_download_xlsx(self, admin_user):
        resp = auth_client(admin_user).get('/api/reporting/stats-export.xlsx/')
        assert resp.status_code == 200
        assert resp['Content-Type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        assert 'Indicateurs' in wb.sheetnames
        assert 'Répartition organisation' in wb.sheetnames

    def test_consultant_forbidden(self, consultant_user):
        resp = auth_client(consultant_user).get('/api/reporting/stats-export.xlsx/')
        assert resp.status_code == 403

    def test_export_is_logged_in_audit(self, admin_user):
        auth_client(admin_user).get('/api/reporting/stats-export.xlsx/')
        assert AuditLog.objects.filter(action=AuditLog.Action.EXPORT, details__type='statistiques').exists()
